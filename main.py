import tkinter as tk
from PIL import Image, ImageTk
from calculate1 import *
from calculate2 import *
import openpyxl
from tkinter import ttk


workbook = openpyxl.load_workbook("record.xlsx")
worksheet = workbook["Sheet1"]


def get_image(filename, width, height):
    im = Image.open(filename).resize((width, height))
    return ImageTk.PhotoImage(im)


def is_scalar(str):
    try:
        float(str)
    except ValueError:
        return False
    else:
        return True


def bt_browse_click():
    lb_welcome.place_forget()
    lb_hint.place_forget()
    lb1.place_forget()
    lb2.place_forget()
    bt1.place_forget()
    bt2.place_forget()
    bt_browse.place_forget()
    global tree, bar, back5, delete
    bar = tk.Scrollbar(win, orient='vertical')
    tree = ttk.Treeview(win, show="headings", columns=('电路类型', 'R1(Ω)', 'R2(Ω)', 'R3(Ω)', 'R4(Ω)', 'U(V)', 'I(A)'),
                        yscrollcommand=bar.set)
    bar['command'] = tree.yview
    bar.place(x=598, y=71, height=249)
    tree.column('电路类型', width=70, anchor='center')
    tree.column('R1(Ω)', width=70, anchor='center')
    tree.column('R2(Ω)', width=70, anchor='center')
    tree.column('R3(Ω)', width=70, anchor='center')
    tree.column('R4(Ω)', width=70, anchor='center')
    tree.column('U(V)', width=70, anchor='center')
    tree.column('I(A)', width=70, anchor='center')
    tree.heading('电路类型', text='电路类型')
    tree.heading('R1(Ω)', text='R1(Ω)')
    tree.heading('R2(Ω)', text='R2(Ω)')
    tree.heading('R3(Ω)', text='R3(Ω)')
    tree.heading('R4(Ω)', text='R4(Ω)')
    tree.heading('U(V)', text='U(V)')
    tree.heading('I(A)', text='I(A)')
    tree.place(x=105, y=70, height=250)
    for i in range(2, worksheet.max_row+1):
        tree.insert('', i, values=(worksheet.cell(row=i, column=1).value,
                                   worksheet.cell(row=i, column=2).value,
                                   worksheet.cell(row=i, column=3).value,
                                   worksheet.cell(row=i, column=4).value,
                                   worksheet.cell(row=i, column=5).value,
                                   worksheet.cell(row=i, column=6).value,
                                   worksheet.cell(row=i, column=7).value))

    delete = tk.Button(win, image=dele, command=delete_click)
    delete.place(x=100, y=380)

    back5 = tk.Button(win, image=back, command=back5_click)
    back5.place(x=490, y=380)


def delete_click():
    rows = worksheet.max_row
    worksheet.delete_rows(2, rows)
    workbook.save('record.xlsx')
    for row in tree.get_children():
        tree.delete(row)
    tree.insert('', 1, values=("", "", "", "", "", "", ""))


def back5_click():
    tree.place_forget()
    bar.place_forget()
    back5.place_forget()
    delete.place_forget()
    global lb_welcome, lb_hint, lb1, lb2, bt1, bt2, bt_browse
    lb_welcome = tk.Label(win, text="欢迎使用通用电流计算器！", bg="black", fg="white", font=('宋体', 30, 'bold'))
    lb_welcome.place(x=100, y=30, width=500)

    lb_hint = tk.Label(win, text="请选择您要计算的电路拓扑：", bg="black", fg="white", font=('宋体', 15, 'bold'))
    lb_hint.place(x=100, y=90)

    lb1 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo1)
    lb1.place(x=80, y=170, width=220, height=160)

    lb2 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo2)
    lb2.place(x=410, y=170, width=220, height=160)

    bt1 = tk.Button(win, image=bt01, command=bt1_click)
    bt1.place(x=135, y=380)

    bt2 = tk.Button(win, image=bt02, command=bt2_click)
    bt2.place(x=472, y=380)

    bt_browse = tk.Button(win, image=bt03, command=bt_browse_click)
    bt_browse.place(x=272, y=390)


# 电路一的函数
def bt1_click():
    lb_welcome.place_forget()
    lb_hint.place_forget()
    lb1.place_forget()
    lb2.place_forget()
    bt1.place_forget()
    bt2.place_forget()
    bt_browse.place_forget()
    global r1, r2, r3, r4, u, lb3, lbr1, lbr2, lbr3, lbr4, lbu, \
        etr1, etr2, etr3, etr4, etu, bt_calculate, back1, lb_warning
    lb3 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo1)
    lb3.place(x=240, y=20, width=220, height=160)
    lb_warning = tk.Label(win, text="", bg="black", fg="red", compound='center', font=17)
    lb_warning.place(x=280, y=410)

    r1, r2, r3, r4, u = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()

    lbr1 = tk.Label(win, text="请输入R1阻值(Ω)：", bg="black", fg="white", font=13)
    lbr1.place(x=180, y=200)
    etr1 = tk.Entry(win, textvariable=r1, font=13)
    etr1.place(x=340, y=200)

    lbr2 = tk.Label(win, text="请输入R2阻值(Ω)：", bg="black", fg="white", font=13)
    lbr2.place(x=180, y=240)
    etr2 = tk.Entry(win, textvariable=r2, font=13)
    etr2.place(x=340, y=240)

    lbr3 = tk.Label(win, text="请输入R3阻值(Ω)：", bg="black", fg="white", font=13)
    lbr3.place(x=180, y=280)
    etr3 = tk.Entry(win, textvariable=r3, font=13)
    etr3.place(x=340, y=280)

    lbr4 = tk.Label(win, text="请输入R4阻值(Ω)：", bg="black", fg="white", font=13)
    lbr4.place(x=180, y=320)
    etr4 = tk.Entry(win, textvariable=r4, font=13)
    etr4.place(x=340, y=320)

    lbu = tk.Label(win, text="请输入电源电压(V)：", bg="black", fg="white", font=13)
    lbu.place(x=180, y=360)
    etu = tk.Entry(win, textvariable=u, font=13)
    etu.place(x=340, y=360)

    bt_calculate = tk.Button(win, image=bt_cal, command=bt1_cal_click)
    bt_calculate.place(x=135, y=410)

    back1 = tk.Button(win, image=back, command=back1_click)
    back1.place(x=472, y=410)


def bt1_cal_click():
    if is_scalar(r1.get()) and is_scalar(r2.get()) \
        and is_scalar(r3.get()) and is_scalar(r4.get()) \
        and is_scalar(u.get()):
        lb_warning.place_forget()
        lb3.place_forget()
        lbr1.place_forget()
        lbr2.place_forget()
        lbr3.place_forget()
        lbr4.place_forget()
        lbu.place_forget()
        etr1.place_forget()
        etr2.place_forget()
        etr3.place_forget()
        etr4.place_forget()
        etu.place_forget()
        bt_calculate.place_forget()
        back1.place_forget()
        global lb_print, back2, lb4
        lb4 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo1)
        lb4.place(x=240, y=20, width=220, height=160)
        R1 = eval(r1.get())
        R2 = eval(r2.get())
        R3 = eval(r3.get())
        R4 = eval(r4.get())
        U = eval(u.get())
        i = calculate1(R1, R2, R3, R4, U)
        rows = worksheet.max_row
        worksheet.cell(rows + 1, 1, "电路1")
        worksheet.cell(rows + 1, 2, r1.get())
        worksheet.cell(rows + 1, 3, r2.get())
        worksheet.cell(rows + 1, 4, r3.get())
        worksheet.cell(rows + 1, 5, r4.get())
        worksheet.cell(rows + 1, 6, u.get())
        worksheet.cell(rows + 1, 7, str(round(i, 3)))
        workbook.save('record.xlsx')
        text = "  R1阻值为 " + r1.get() + "Ω  " + "\n" + "\n" + "  R2阻值为 " + r2.get() + "Ω  " + "\n" + "\n" + \
               "  R3阻值为 " + r3.get() + "Ω  " + "\n" + "\n" + "  R4阻值为 " + r4.get() + "Ω  " + "\n" + "\n" + \
               "  电源电压U为 " + u.get() + "V  " + "\n" + "\n" + "\n" + "  电流I为 " + str(round(i, 3)) + "A  "
        lb_print = tk.Label(win, text=text, bg="black", fg="white", font=13)
        lb_print.place(x=273, y=205)
        back2 = tk.Button(win, image=back, command=back2_click)
        back2.place(x=472, y=410)
    else:
        lb_warning["text"] = "请输入有效数据！"
        etr1.delete(0, 10000)
        etr2.delete(0, 10000)
        etr3.delete(0, 10000)
        etr4.delete(0, 10000)
        etu.delete(0, 10000)


def back2_click():
    lb_print.place_forget()
    back2.place_forget()
    lb4.place_forget()
    global r1, r2, r3, r4, u, lb3, lbr1, lbr2, lbr3, lbr4, lbu, \
        etr1, etr2, etr3, etr4, etu, bt_calculate, back1, lb_warning
    lb3 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo1)
    lb3.place(x=240, y=20, width=220, height=160)
    lb_warning = tk.Label(win, text="", bg="black", fg="red", compound='center', font=17)
    lb_warning.place(x=280, y=410)

    r1, r2, r3, r4, u = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()

    lbr1 = tk.Label(win, text="请输入R1阻值(Ω)：", bg="black", fg="white", font=13)
    lbr1.place(x=180, y=200)
    etr1 = tk.Entry(win, textvariable=r1, font=13)
    etr1.place(x=340, y=200)

    lbr2 = tk.Label(win, text="请输入R2阻值(Ω)：", bg="black", fg="white", font=13)
    lbr2.place(x=180, y=240)
    etr2 = tk.Entry(win, textvariable=r2, font=13)
    etr2.place(x=340, y=240)

    lbr3 = tk.Label(win, text="请输入R3阻值(Ω)：", bg="black", fg="white", font=13)
    lbr3.place(x=180, y=280)
    etr3 = tk.Entry(win, textvariable=r3, font=13)
    etr3.place(x=340, y=280)

    lbr4 = tk.Label(win, text="请输入R4阻值(Ω)：", bg="black", fg="white", font=13)
    lbr4.place(x=180, y=320)
    etr4 = tk.Entry(win, textvariable=r4, font=13)
    etr4.place(x=340, y=320)

    lbu = tk.Label(win, text="请输入电源电压(V)：", bg="black", fg="white", font=13)
    lbu.place(x=180, y=360)
    etu = tk.Entry(win, textvariable=u, font=13)
    etu.place(x=340, y=360)

    bt_calculate = tk.Button(win, image=bt_cal, command=bt1_cal_click)
    bt_calculate.place(x=135, y=410)

    back1 = tk.Button(win, image=back, command=back1_click)
    back1.place(x=472, y=410)


def back1_click():
    lb_warning.place_forget()
    lb3.place_forget()
    lbr1.place_forget()
    lbr2.place_forget()
    lbr3.place_forget()
    lbr4.place_forget()
    lbu.place_forget()
    etr1.place_forget()
    etr2.place_forget()
    etr3.place_forget()
    etr4.place_forget()
    etu.place_forget()
    bt_calculate.place_forget()
    back1.place_forget()
    global lb_welcome, lb_hint, lb1, lb2, bt1, bt2, bt_browse
    lb_welcome = tk.Label(win, text="欢迎使用通用电流计算器！", bg="black", fg="white", font=('宋体', 30, 'bold'))
    lb_welcome.place(x=100, y=30, width=500)

    lb_hint = tk.Label(win, text="请选择您要计算的电路拓扑：", bg="black", fg="white", font=('宋体', 15, 'bold'))
    lb_hint.place(x=100, y=90)

    lb1 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo1)
    lb1.place(x=80, y=170, width=220, height=160)

    lb2 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo2)
    lb2.place(x=410, y=170, width=220, height=160)

    bt1 = tk.Button(win, image=bt01, command=bt1_click)
    bt1.place(x=135, y=380)

    bt2 = tk.Button(win, image=bt02, command=bt2_click)
    bt2.place(x=472, y=380)

    bt_browse = tk.Button(win, image=bt03, command=bt_browse_click)
    bt_browse.place(x=272, y=390)


# 电路二部分的函数
def bt2_click():
    lb_welcome.place_forget()
    lb_hint.place_forget()
    lb1.place_forget()
    lb2.place_forget()
    bt1.place_forget()
    bt2.place_forget()
    bt_browse.place_forget()
    global r1, r2, r3, r4, u, lb3, lbr1, lbr2, lbr3, lbr4, lbu,\
        etr1, etr2, etr3, etr4, etu, bt_calculate, back3, lb_warning
    lb3 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo2)
    lb3.place(x=240, y=20, width=220, height=160)
    lb_warning = tk.Label(win, text="", bg="black", fg="red", compound='center', font=17)
    lb_warning.place(x=280, y=410)

    r1, r2, r3, r4, u = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()

    lbr1 = tk.Label(win, text="请输入R1阻值(Ω)：", bg="black", fg="white", font=13)
    lbr1.place(x=180, y=200)
    etr1 = tk.Entry(win, textvariable=r1, font=13)
    etr1.place(x=340, y=200)

    lbr2 = tk.Label(win, text="请输入R2阻值(Ω)：", bg="black", fg="white", font=13)
    lbr2.place(x=180, y=240)
    etr2 = tk.Entry(win, textvariable=r2, font=13)
    etr2.place(x=340, y=240)

    lbr3 = tk.Label(win, text="请输入R3阻值(Ω)：", bg="black", fg="white", font=13)
    lbr3.place(x=180, y=280)
    etr3 = tk.Entry(win, textvariable=r3, font=13)
    etr3.place(x=340, y=280)

    lbr4 = tk.Label(win, text="请输入R4阻值(Ω)：", bg="black", fg="white", font=13)
    lbr4.place(x=180, y=320)
    etr4 = tk.Entry(win, textvariable=r4, font=13)
    etr4.place(x=340, y=320)

    lbu = tk.Label(win, text="请输入电源电压(V)：", bg="black", fg="white", font=13)
    lbu.place(x=180, y=360)
    etu = tk.Entry(win, textvariable=u, font=13)
    etu.place(x=340, y=360)

    bt_calculate = tk.Button(win, image=bt_cal, command=bt2_cal_click)
    bt_calculate.place(x=135, y=410)

    back3 = tk.Button(win, image=back, command=back3_click)
    back3.place(x=472, y=410)


def bt2_cal_click():
    if is_scalar(r1.get()) and is_scalar(r2.get()) \
        and is_scalar(r3.get()) and is_scalar(r4.get()) \
        and is_scalar(u.get()):
        lb_warning.place_forget()
        lb3.place_forget()
        lbr1.place_forget()
        lbr2.place_forget()
        lbr3.place_forget()
        lbr4.place_forget()
        lbu.place_forget()
        etr1.place_forget()
        etr2.place_forget()
        etr3.place_forget()
        etr4.place_forget()
        etu.place_forget()
        bt_calculate.place_forget()
        back3.place_forget()
        global lb_print, back4, lb4
        lb4 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo2)
        lb4.place(x=240, y=20, width=220, height=160)
        R1 = eval(r1.get())
        R2 = eval(r2.get())
        R3 = eval(r3.get())
        R4 = eval(r4.get())
        U = eval(u.get())
        i = calculate2(R1, R2, R3, R4, U)
        rows = worksheet.max_row
        worksheet.cell(rows + 1, 1, "电路2")
        worksheet.cell(rows + 1, 2, r1.get())
        worksheet.cell(rows + 1, 3, r2.get())
        worksheet.cell(rows + 1, 4, r3.get())
        worksheet.cell(rows + 1, 5, r4.get())
        worksheet.cell(rows + 1, 6, u.get())
        worksheet.cell(rows + 1, 7, str(round(i, 3)))
        workbook.save('record.xlsx')
        text = "  R1阻值为 " + r1.get() + "Ω  " + "\n" + "\n" + "  R2阻值为 " + r2.get() + "Ω  " + "\n" + "\n" + \
               "  R3阻值为 " + r3.get() + "Ω  " + "\n" + "\n" + "  R4阻值为 " + r4.get() + "Ω  " + "\n" + "\n" + \
               "  电源电压U为 " + u.get() + "V  " + "\n" + "\n" + "\n" + "  电流I为 " + str(round(i, 3)) + "A  "
        lb_print = tk.Label(win, text=text, bg="black", fg="white", font=13)
        lb_print.place(x=273, y=205)
        back4 = tk.Button(win, image=back, command=back4_click)
        back4.place(x=472, y=410)
    else:
        lb_warning["text"] = "请输入有效数据！"
        etr1.delete(0, 10000)
        etr2.delete(0, 10000)
        etr3.delete(0, 10000)
        etr4.delete(0, 10000)
        etu.delete(0, 10000)


def back4_click():
    lb4.place_forget()
    lb_print.place_forget()
    back4.place_forget()
    global r1, r2, r3, r4, u, lb3, lbr1, lbr2, lbr3, lbr4, lbu, \
        etr1, etr2, etr3, etr4, etu, bt_calculate, back3, lb_warning
    lb3 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo2)
    lb3.place(x=240, y=20, width=220, height=160)
    lb_warning = tk.Label(win, text="", bg="black", fg="red", compound='center', font=17)
    lb_warning.place(x=280, y=410)

    r1, r2, r3, r4, u = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()

    lbr1 = tk.Label(win, text="请输入R1阻值(Ω)：", bg="black", fg="white", font=13)
    lbr1.place(x=180, y=200)
    etr1 = tk.Entry(win, textvariable=r1, font=13)
    etr1.place(x=340, y=200)

    lbr2 = tk.Label(win, text="请输入R2阻值(Ω)：", bg="black", fg="white", font=13)
    lbr2.place(x=180, y=240)
    etr2 = tk.Entry(win, textvariable=r2, font=13)
    etr2.place(x=340, y=240)

    lbr3 = tk.Label(win, text="请输入R3阻值(Ω)：", bg="black", fg="white", font=13)
    lbr3.place(x=180, y=280)
    etr3 = tk.Entry(win, textvariable=r3, font=13)
    etr3.place(x=340, y=280)

    lbr4 = tk.Label(win, text="请输入R4阻值(Ω)：", bg="black", fg="white", font=13)
    lbr4.place(x=180, y=320)
    etr4 = tk.Entry(win, textvariable=r4, font=13)
    etr4.place(x=340, y=320)

    lbu = tk.Label(win, text="请输入电源电压(V)：", bg="black", fg="white", font=13)
    lbu.place(x=180, y=360)
    etu = tk.Entry(win, textvariable=u, font=13)
    etu.place(x=340, y=360)

    bt_calculate = tk.Button(win, image=bt_cal, command=bt2_cal_click)
    bt_calculate.place(x=135, y=410)

    back3 = tk.Button(win, image=back, command=back3_click)
    back3.place(x=472, y=410)


def back3_click():
    lb_warning.place_forget()
    lb3.place_forget()
    lbr1.place_forget()
    lbr2.place_forget()
    lbr3.place_forget()
    lbr4.place_forget()
    lbu.place_forget()
    etr1.place_forget()
    etr2.place_forget()
    etr3.place_forget()
    etr4.place_forget()
    etu.place_forget()
    bt_calculate.place_forget()
    back3.place_forget()
    global lb_welcome, lb_hint, lb1, lb2, bt1, bt2, bt_browse
    lb_welcome = tk.Label(win, text="欢迎使用通用电流计算器！", bg="black", fg="white", font=('宋体', 30, 'bold'))
    lb_welcome.place(x=100, y=30, width=500)

    lb_hint = tk.Label(win, text="请选择您要计算的电路拓扑：", bg="black", fg="white", font=('宋体', 15, 'bold'))
    lb_hint.place(x=100, y=90)

    lb1 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo1)
    lb1.place(x=80, y=170, width=220, height=160)

    lb2 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo2)
    lb2.place(x=410, y=170, width=220, height=160)

    bt1 = tk.Button(win, image=bt01, command=bt1_click)
    bt1.place(x=135, y=380)

    bt2 = tk.Button(win, image=bt02, command=bt2_click)
    bt2.place(x=472, y=380)

    bt_browse = tk.Button(win, image=bt03, command=bt_browse_click)
    bt_browse.place(x=272, y=390)


win = tk.Tk()
win.title("通用电流计算器")
win.geometry("700x500")
win.attributes("-toolwindow", 2)
win.resizable(0, 0)
im = get_image("1.png", 700, 500)
canvas_win = tk.Canvas(win, width=700, height=500)
canvas_win.create_image(350, 250, image=im)
canvas_win.pack()
bt_cal = get_image("cal.png", 87, 60)
back = get_image("back.png", 87, 60)
dele = get_image("del.png", 214, 60)

lb_welcome = tk.Label(win, text="欢迎使用通用电流计算器！", bg="black", fg="white", font=('宋体', 30, 'bold'))
lb_welcome.place(x=100, y=30, width=500)

lb_hint = tk.Label(win, text="请选择您要计算的电路拓扑：", bg="black", fg="white", font=('宋体', 15, 'bold'))
lb_hint.place(x=100, y=90)

photo1 = get_image("loop1.png", 210, 150)
lb1 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo1)
lb1.place(x=80, y=170, width=220, height=160)

photo2 = get_image("loop2.png", 210, 150)
lb2 = tk.Label(win, bg="DodgerBlue2", compound='center', image=photo2)
lb2.place(x=410, y=170, width=220, height=160)

bt01 = get_image("bt1.png", 87, 60)
bt1 = tk.Button(win, image=bt01, command=bt1_click)
bt1.place(x=135, y=380)

bt02 = get_image("bt2.png", 87, 60)
bt2 = tk.Button(win, image=bt02, command=bt2_click)
bt2.place(x=472, y=380)

bt03 = get_image("Browse.png", 150, 42)
bt_browse = tk.Button(win, image=bt03, command=bt_browse_click)
bt_browse.place(x=272, y=390)

win.mainloop()
